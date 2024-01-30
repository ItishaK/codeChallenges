import openpyxl
from orderedset import OrderedSet

path = "DDL_File.xlsx"

wb_obj = openpyxl.load_workbook("DDL_File.xlsx")
sheet_obj = wb_obj.active
tbl_header = sheet_obj.cell(row=1, column=1)
col_header = sheet_obj.cell(row=1, column=2)

i = 2
tbl_nm_arr = []
tbl_nm_set = {}
temp = sheet_obj.cell(row=2, column=1)
while temp is not None:
    tbl_nm_arr.append(sheet_obj.cell(row=i, column=1).value)
    i += 1
    temp = sheet_obj.cell(row=i, column=1).value

tbl_nm_set = OrderedSet(tbl_nm_arr)

rw = 2
temp_tbl_nm = sheet_obj.cell(row=rw, column=1).value

op_file = open("finalDDL.txt", "w")
for j in tbl_nm_set:
    constraint_nm = []
    constraint_typ = []
    ref_col = []
    ref_col_tbl = []
    pk = ""
    str0 = "CREATE TABLE "+j+" ("
    op_file.write(str0)
    op_file.write("\n")
    while j == temp_tbl_nm:
        col_nm = sheet_obj.cell(row=rw, column=2).value
        dta_typ = sheet_obj.cell(row=rw, column=3).value
        constraint_nm.append(sheet_obj.cell(row=rw, column=4).value)
        constraint_typ.append(sheet_obj.cell(row=rw, column=5).value)
        ref_col.append(sheet_obj.cell(row=rw, column=6).value)
        ref_col_tbl.append(sheet_obj.cell(row=rw, column=7).value)
        str2 = col_nm+" "+dta_typ
        rw += 1
        temp_tbl_nm = sheet_obj.cell(row=rw, column=1).value

        if temp_tbl_nm == j:
            op_file.write(str2+",")
            op_file.write("\n")
        else:
            for k in constraint_nm:
                if k is not None:
                    print('constraint_nm: ' + k)
                    loc = constraint_nm.index(k)
                    print('constraint_typ: ' + constraint_typ[loc])
                    if constraint_typ[loc] == 'PK':
                        pk = "CONSTRAINT " + constraint_nm[loc] + " PRIMARY KEY (" + ref_col[loc] + ")"
                        op_file.write(str2+",")
                        op_file.write("\n")
                        op_file.write(pk)
            if pk == "":
                op_file.write(str2)
                op_file.write("\n")

    op_file.write(");")
    op_file.write("\n\n")

op_file.close()
